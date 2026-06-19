"""
Compare the shape preservation (GT vs extracted beat-amplitude correlation, per class)
across several extraction models -> pick the best SUBSTRATE for the adapter.
Reuses the saved extractions (inferred_<v>/) when present; otherwise extracts.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
import pathsetup  # noqa
import numpy as np
import torch
torch.set_num_threads(1)
from diag_info_ceiling import detect_peaks, amp_at, load_sig, load_mask, CLASS_NAME
import pipeline_registry as R

from config import NPY, FECG_ROOT as OUT_BASE, RESULTS_DIR as RES, XLSX
MODELS = ['v1', 'v5', 'v6', 'v15', 'v16', 'v17']
EPS = 1e-6


def get_ext(v, name, model):
    saved = os.path.join(OUT_BASE, f'inferred_{v}', f'{name}.npy')
    if os.path.exists(saved):
        return np.load(saved)
    return R.infer(v, model, load_sig('mixture', name))


def diagnose(v):
    signals = sorted(os.listdir(os.path.join(NPY, 'mixture')))
    sample = signals[::7]   # same set as the v1 diagnostic
    model = None
    if not os.path.exists(os.path.join(OUT_BASE, f'inferred_{v}')):
        model = R.load_extractor(v)
    glob_r = []
    seg = {c: [] for c in (1, 2, 3)}        # shape correlation (movement only)
    seg_w = {c: [] for c in (1, 2, 3)}
    nm_gt, nm_ext = [], []                   # no-move std (normalized to the per-segment median)
    for name in sample:
        gt = load_sig('signals', name)
        ext = get_ext(v, name, model)
        mc = load_mask('mc_masks', name).astype(int)
        N = min(gt.shape[1], ext.shape[1], len(mc))
        for ch in range(6):
            pk = detect_peaks(ext[ch, :N])
            if len(pk) < 20:
                continue
            ag = amp_at(gt[ch, :N], pk); ae = amp_at(ext[ch, :N], pk)
            if ag.std() > 1e-6 and ae.std() > 1e-6:
                glob_r.append(np.corrcoef(ag, ae)[0, 1])
            cls = mc[pk]; b = 0
            while b < len(pk):
                e = b
                while e + 1 < len(pk) and cls[e + 1] == cls[b]: e += 1
                run = slice(b, e + 1); L = e + 1 - b; c = int(cls[b])
                if L >= 6:
                    if c == 0:   # no-move: std normalized to the median (how flat it is)
                        nm_gt.append(ag[run].std() / (np.median(ag[run]) + EPS))
                        nm_ext.append(ae[run].std() / (np.median(ae[run]) + EPS))
                    elif ag[run].std() > 1e-6 and ae[run].std() > 1e-6:
                        seg[c].append(float(np.corrcoef(ag[run], ae[run])[0, 1])); seg_w[c].append(L)
                b = e + 1
    gr = float(np.nanmean(glob_r))
    per = {}
    for c in (1, 2, 3):
        if seg[c]:
            r = np.array(seg[c]); w = np.array(seg_w[c])
            per[c] = float(np.nansum(r * w) / np.nansum(w[~np.isnan(r)])) if np.any(~np.isnan(r)) else float('nan')
        else:
            per[c] = float('nan')
    per['nm_gt_std'] = float(np.nanmean(nm_gt)) if nm_gt else float('nan')
    per['nm_ext_std'] = float(np.nanmean(nm_ext)) if nm_ext else float('nan')
    return gr, per


def main():
    print('Shape correlation (high r = good) for linear/spline/helix; no-move = std (extracted close to GT = good)\n', flush=True)
    print(f"{'model':<6}{'global_r':>9}{'linear':>8}{'spline':>8}{'helix':>8}{'  | nm_std GT->ext':>20}", flush=True)
    print('-' * 62, flush=True)
    results = {}
    for v in MODELS:
        gr, per = diagnose(v)
        results[v] = (gr, per)
        print(f"{v:<6}{gr:>9.3f}{per[1]:>8.3f}{per[2]:>8.3f}{per[3]:>8.3f}"
              f"   {per['nm_gt_std']:.3f}->{per['nm_ext_std']:.3f}", flush=True)
    print('\nSubstrate score (how good it is for the adapter):', flush=True)
    print('  - move_mean = mean correlation over linear/spline/helix (higher = shape preserved)', flush=True)
    print('  - excess_nm = std_ext/std_gt in no-move (closer to 1 = less spurious noise to flatten)', flush=True)
    best = None
    for v in MODELS:
        gr, per = results[v]
        mv = float(np.nanmean([per[1], per[2], per[3]]))
        excess = per['nm_ext_std'] / (per['nm_gt_std'] + EPS)
        print(f"  {v}: move_mean={mv:.3f}  excess_nm={excess:.2f}x", flush=True)
        score = mv  # priority: shape preserved
        if best is None or score > best[1]:
            best = (v, score)
    print(f"\n=> best substrate by shape preservation: {best[0]} (move_mean={best[1]:.3f})", flush=True)
    save_all(results)


def save_all(results):
    """results[v] = (global_r, {1,2,3,nm_gt_std,nm_ext_std}). Saves JSON + xlsx + tex."""
    import json
    out = {v: {'global_r': gr, 'linear': per[1], 'spline': per[2], 'helix': per[3],
               'nm_gt_std': per['nm_gt_std'], 'nm_ext_std': per['nm_ext_std']}
           for v, (gr, per) in results.items()}
    jpath = os.path.join(RES, 'substrate_shape_preservation.json')
    json.dump(out, open(jpath, 'w'), indent=2)
    print(f'JSON: {jpath}', flush=True)

    # xlsx
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.load_workbook(XLSX); SH = 'Substrate shape preserv'
    if SH in wb.sheetnames: del wb[SH]
    ws = wb.create_sheet(SH)
    bold = Font(bold=True); ital = Font(italic=True, size=9); ctr = Alignment(horizontal='center')
    thin = Side(style='thin', color='BBBBBB'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    L = [('Preservation of the movement signature after extraction (GT vs extracted R-peak amplitude correlation, per class)', bold),
         ('linear/spline/helix: shape correlation within segments (1=perfectly preserved, 0=lost). 18 test signals, mean over channels.', ital),
         ('no-move: normalized amplitude std (GT->extracted); extracted >> GT = spurious variation added = over-detection.', ital),
         ('Ideal substrate for the adapter: high linear/spline/helix + no-move ext close to GT.', ital), ('', None)]
    r = 1
    for txt, f in L:
        cc = ws.cell(r, 1, txt)
        if f: cc.font = f
        r += 1
    for j, h in enumerate(['Model', 'global r', 'linear', 'spline', 'helix', 'no-move std GT', 'no-move std ext']):
        cc = ws.cell(r, 1+j, h); cc.font = bold; cc.alignment = ctr; cc.fill = PatternFill('solid', fgColor='D9E1F2'); cc.border = bd
    r += 1
    for v in MODELS:
        gr, per = results[v]
        for j, val in enumerate([v, round(gr, 3), round(per[1], 3), round(per[2], 3), round(per[3], 3),
                                 round(per['nm_gt_std'], 3), round(per['nm_ext_std'], 3)]):
            cc = ws.cell(r, 1+j, val); cc.border = bd
            if j >= 1: cc.alignment = ctr
        r += 1
    for col in 'ABCDEFG': ws.column_dimensions[col].width = 14
    wb.save(XLSX)
    print(f'xlsx: sheet "{SH}"', flush=True)

    # tex
    tpath = os.path.join(RES, 'substrate_shape_preservation.tex')
    with open(tpath, 'w') as f:
        f.write('% Preservation of the movement signature per extraction model (per-class amplitude-shape correlation GT vs extracted).\n')
        f.write('\\begin{table}[h]\\centering\n')
        f.write('\\caption{Preservation of the movement signature after fetal ECG extraction: Pearson correlation '
                'between the ground-truth and extracted R-peak amplitude sequences, per movement type, with the '
                'normalized amplitude spread during no-movement (lower extracted spread is better).}\n')
        f.write('\\label{tab:substrate_preservation}\n')
        f.write('\\begin{tabular}{lccccc}\n\\hline\n')
        f.write('Model & Linear & Spline & Helix & No-move spread (GT) & No-move spread (ext.) \\\\\n\\hline\n')
        for v in MODELS:
            gr, per = results[v]
            f.write(f'{v} & {per[1]:.3f} & {per[2]:.3f} & {per[3]:.3f} & {per["nm_gt_std"]:.3f} & {per["nm_ext_std"]:.3f} \\\\\n')
        f.write('\\hline\n\\end{tabular}\n\\end{table}\n')
    print(f'tex: {tpath}', flush=True)


if __name__ == '__main__':
    main()
