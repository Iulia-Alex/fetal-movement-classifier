"""
v1 adapter trained on DB_1 (the TRAIN base, .mat) — the correct variant compared to the POC
(which trained on a split of the test set). GT used ONLY at training on DB_1; at eval
(the 9 held-out signals from Final_Test_DB) the adapter does NOT see GT.
Same mechanism: beat-amplitude regression extracted->GT, re-injection as gain, M15.
"""
import sys, os, json, glob, time
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
import pathsetup  # noqa
import numpy as np
import torch, torch.nn as nn
torch.set_num_threads(1)
from movement_dataset import _load_mat
from diag_info_ceiling import detect_peaks, amp_at, load_sig, load_mask
from build_adapter import Adapter, windows, correct_signal, classify_m15, snr_cat, K, EPS
import pipeline_registry as R
import pretrained_clf as E
from sklearn.metrics import accuracy_score, f1_score

from config import DB1_LONG, EVAL_DIR, NPY, XLSX, RESULTS_DIR as _RES
OUT = _RES + '/adapter_v1_db1.json'
N_TRAIN_MAT = 60


def build_pairs(mats, model):
    X, Y = [], []
    t0 = time.time()
    for i, p in enumerate(mats):
        try:
            mixture, fecg = _load_mat(p)
        except Exception as e:
            print(f'  skip {os.path.basename(p)}: {e}', flush=True); continue
        ext = R.infer('v1', model, mixture)
        M = min(ext.shape[1], fecg.shape[1])
        for ch in range(6):
            pk = detect_peaks(ext[ch, :M])
            if len(pk) < K + 5:
                continue
            ae = amp_at(ext[ch, :M], pk); ag = amp_at(fecg[ch, :M], pk)
            e = ae / (np.median(ae) + EPS); g = ag / (np.median(ag) + EPS)
            X.append(windows(e)); Y.append(g.astype(np.float32))
        if (i + 1) % 20 == 0:
            print(f'  [{i+1}/{len(mats)}] {(time.time()-t0)/60:.0f}min', flush=True)
    return np.concatenate(X), np.concatenate(Y)


def train(X, Y, epochs=40):
    Xt = torch.from_numpy(X); Yt = torch.from_numpy(Y)
    m = Adapter(); opt = torch.optim.Adam(m.parameters(), lr=1e-3); lossf = nn.MSELoss()
    n = len(Xt); bs = 512
    for ep in range(epochs):
        perm = torch.randperm(n)
        for i in range(0, n, bs):
            idx = perm[i:i+bs]; opt.zero_grad()
            lossf(m(Xt[idx]), Yt[idx]).backward(); opt.step()
    return m.eval()


def eval_heldout(adapter, model, clf15):
    sigs = sorted(os.listdir(os.path.join(NPY, 'mixture')))
    v1done = set(json.load(open(os.path.join(EVAL_DIR, 'v1.json'))).keys())
    by = {'easy': [], 'medium': [], 'difficult': []}
    for s in sigs:
        if s in v1done: by[snr_cat(s)].append(s)
    eval_sigs = [by[c][i] for c in by for i in range(3)]   # same set as the POC
    gt_json = json.load(open(os.path.join(EVAL_DIR, 'GT.json')))
    v1_json = json.load(open(os.path.join(EVAL_DIR, 'v1.json')))
    rows = []
    for nm in eval_sigs:
        gt = load_sig('signals', nm); ext = R.infer('v1', model, load_sig('mixture', nm))
        mc = load_mask('mc_masks', nm).astype(int)
        N = min(gt.shape[1], ext.shape[1], len(mc)); mc = mc[:N]
        acc_cor = []
        for ch in range(6):
            pk = detect_peaks(ext[ch, :N])
            if len(pk) < K + 5: continue
            e = amp_at(ext[ch, :N], pk); e = e / (np.median(e) + EPS)
            xc = correct_signal(ext[ch, :N], pk, e, adapter)
            acc_cor.append(accuracy_score(mc, classify_m15(xc, clf15)[:N]))
        gt_acc = float(np.mean([c['m15_mc_acc'] for c in gt_json[nm]['channels']]))
        v1_acc = float(np.mean([c['m15_mc_acc'] for c in v1_json[nm]['channels']]))
        cor = float(np.mean(acc_cor))
        rows.append((nm, snr_cat(nm), gt_acc, v1_acc, cor))
        print(f'  {nm.split("_SNR")[0]:<12} {snr_cat(nm):<10} GT={gt_acc:.3f}  v1={v1_acc:.3f}  v1+adapter(DB1)={cor:.3f}', flush=True)
    return rows


def main():
    model = R.load_extractor('v1')
    mats = sorted(glob.glob(os.path.join(DB1_LONG, '*.mat')))
    step = max(1, len(mats) // N_TRAIN_MAT)
    train_mats = mats[::step][:N_TRAIN_MAT]
    print(f'Train on {len(train_mats)} .mat from DB_1 (Long). Building pairs...', flush=True)
    t = time.time()
    X, Y = build_pairs(train_mats, model)
    print(f'  pairs: {len(X)} ({(time.time()-t)/60:.0f}min)', flush=True)
    adapter = train(X, Y)
    print('  adapter trained.', flush=True)

    clf15 = E.load_model(15)
    rows = eval_heldout(adapter, model, clf15)
    g = np.mean([r[2] for r in rows]); v = np.mean([r[3] for r in rows]); c = np.mean([r[4] for r in rows])
    print(f'\n=== DOWNSTREAM M15 (DB_1-trained, 9 held-out) ===')
    print(f'  GT={g:.3f}  v1(uncor)={v:.3f}  v1+adapter(DB1)={c:.3f}  | delta={c-v:+.3f}')
    print(f'  (POC, trained on test split: v1+adapter=0.485, delta=+0.123)')
    json.dump({'rows': rows, 'mean': {'GT': g, 'v1': v, 'v1_adapter_db1': c}}, open(OUT, 'w'), indent=2)
    write_xlsx(rows)


def write_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.load_workbook(XLSX); SH = 'Adapter v1 (DB_1)'
    if SH in wb.sheetnames: del wb[SH]
    ws = wb.create_sheet(SH)
    bold = Font(bold=True); ital = Font(italic=True, size=9); ctr = Alignment(horizontal='center')
    thin = Side(style='thin', color='BBBBBB'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    L = [('v1 ADAPTER -> M15, trained on DB_1 (the TRAIN base), eval on 9 held-out signals from test', bold),
         ('GT used ONLY at training on DB_1; at eval the adapter sees only the extracted signal (no GT) -> methodologically correct.', ital),
         ('GT/v1 read from the big job; v1+adapter(DB1) computed here. Multiclass acc, 4 classes, mean over 6 channels.', ital), ('', None)]
    r = 1
    for txt, f in L:
        cc = ws.cell(r, 1, txt)
        if f: cc.font = f
        r += 1
    for j, h in enumerate(['Signal', 'Category', 'GT (ceiling)', 'v1 (uncor)', 'v1+adapter(DB1)', 'delta']):
        cc = ws.cell(r, 1+j, h); cc.font = bold; cc.alignment = ctr; cc.fill = PatternFill('solid', fgColor='E2EFDA'); cc.border = bd
    r += 1
    for nm, cat, g, v, c in rows:
        for j, val in enumerate([nm.split('_SNR')[0], cat, round(g, 3), round(v, 3), round(c, 3), round(c-v, 3)]):
            cc = ws.cell(r, 1+j, val); cc.border = bd
            if j >= 2: cc.alignment = ctr
        r += 1
    gm = np.mean([x[2] for x in rows]); vm = np.mean([x[3] for x in rows]); cm = np.mean([x[4] for x in rows])
    for j, val in enumerate(['MEAN', '', round(gm, 3), round(vm, 3), round(cm, 3), round(cm-vm, 3)]):
        cc = ws.cell(r, 1+j, val); cc.font = bold; cc.border = bd
        if j >= 2: cc.alignment = ctr
    ws.column_dimensions['A'].width = 16
    for col in 'BCDEF': ws.column_dimensions[col].width = 15
    wb.save(XLSX)
    print(f'\nsheet "{SH}" written to {XLSX}')


if __name__ == '__main__':
    main()
