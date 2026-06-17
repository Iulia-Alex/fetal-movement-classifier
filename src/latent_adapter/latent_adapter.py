"""
ADAPTER LATENT (transfer in spatiul din mijlocul retelei) — directiva profei.
In loc sa corectam semnalul DECODAT (build_adapter / train_adapter_db1), corectam
reprezentarea LATENTA la bottleneck-ul clasificatorului M15 (AttUNet1D, multiclass).

  encoder(features(extras)) -> b_ext (256,240) + skip-uri s1..s4 (necorectate)
  A(b_ext) -> b_corr            (MLP per-pozitie pe cele 256 canale, rezidual)
  decoder INGHETAT (b_corr, s1..s4) -> logits 4 clase

Loss = CrossEntropy prin DECODER-UL INGHETAT vs masca de categorii (LABEL-ONLY:
zero semnal GT, nici la antrenare). A invata ce bottleneck face decoderul sa
clasifice corect DAT FIIND skip-urile extrase reale -> compenseaza scurgerea prin
skip-uri (att-gates), in loc sa o ignore (ce ar face un MSE catre b_gt).

Antrenare pe DB_1 (baza de train, .mat cu category_mask); eval pe held-out din test.
"""
import sys, os, glob, time, json
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
import pathsetup  # noqa
import numpy as np
import torch, torch.nn as nn
import scipy.io as sio
torch.set_num_threads(int(os.environ.get('LA_THREADS', '6')))
from diag_info_ceiling import load_sig, load_mask
from build_adapter import snr_cat
import pipeline_registry as R
import pretrained_clf as E
from sklearn.metrics import accuracy_score, f1_score

from config import DB1_LONG, EVAL_DIR, NPY, XLSX, RESULTS_DIR as _RES
OUT  = _RES + '/adapter_latent_v1.json'
CKPT = _RES + '/adapter_latent_v1.pt'
WIN = E.WIN  # 3840
N_CLASS = 4


# ── M15 encoder/decoder split (verified: decode(*encode(x)) == M15(x)) ──────────
def encode(net, x):
    s1 = net.enc1(x); s2 = net.enc2(net.pool(s1))
    s3 = net.enc3(net.pool(s2)); s4 = net.enc4(net.pool(s3))
    b = net.bottleneck(net.pool(s4))
    if net.self_attn is not None:
        b = net.self_attn(b)
    return b, (s1, s2, s3, s4)


def decode(net, b, skips):
    s1, s2, s3, s4 = skips
    g = net.up(b);  d4 = net.dec4(torch.cat([g, net.att4(g, s4)], 1))
    g = net.up(d4); d3 = net.dec3(torch.cat([g, net.att3(g, s3)], 1))
    g = net.up(d3); d2 = net.dec2(torch.cat([g, net.att2(g, s2)], 1))
    g = net.up(d2); d1 = net.dec1(torch.cat([g, net.att1(g, s1)], 1))
    return net.final(d1)


# ── latent adapter: per-position MLP on bottleneck (256ch), residual ───────────
class LatentAdapter(nn.Module):
    def __init__(self, ch=256, hidden=256):
        super().__init__()
        self.net = nn.Sequential(
            nn.Conv1d(ch, hidden, 1), nn.GroupNorm(8, hidden), nn.ReLU(),
            nn.Conv1d(hidden, ch, 1))
        nn.init.zeros_(self.net[-1].weight)        # start ca identitate (b_corr=b)
        nn.init.zeros_(self.net[-1].bias)

    def forward(self, b):
        return b + self.net(b)


def load_mat_mix_cat(path):
    mat = sio.loadmat(path)
    mix = mat['out']['mixture'][0][0].astype(np.float32)         # (6,600000)
    cat = mat['out']['category_mask'][0][0].ravel().astype(np.int64)  # (600000,)
    del mat
    return mix, cat


# ── build cache (b_ext, skips, label) per window ──────────────────────────────
def build_cache(mats, model, net, n_windows, seed=0):
    rng = np.random.default_rng(seed)
    per_ch = max(1, n_windows // (len(mats) * 6))
    B, S = [], [[], [], [], []]; Y = []
    t0 = time.time()
    for fi, p in enumerate(mats):
        try:
            mix, cat = load_mat_mix_cat(p)
        except Exception as e:
            print(f'  skip {os.path.basename(p)}: {e}', flush=True); continue
        ext = R.infer('v1', model, mix)               # (6,N)
        N = min(ext.shape[1], len(cat))
        max_start = N - WIN
        for ch in range(6):
            starts = rng.integers(0, max_start, size=per_ch)
            feats = np.stack([E.features_5(ext[ch, s0:s0 + WIN]) for s0 in starts])  # (k,5,WIN)
            with torch.no_grad():
                b, skips = encode(net, torch.from_numpy(feats))
            B.append(b.half().numpy())
            for j in range(4):
                S[j].append(skips[j].half().numpy())
            Y.append(np.stack([cat[s0:s0 + WIN] for s0 in starts]).astype(np.int8))
        if (fi + 1) % 5 == 0:
            print(f'  [{fi+1}/{len(mats)}] {(time.time()-t0)/60:.1f}min', flush=True)
    B = np.concatenate(B); S = [np.concatenate(s) for s in S]; Y = np.concatenate(Y)
    print(f'  cache: {len(B)} windows, ~{(B.nbytes+sum(s.nbytes for s in S))/1e9:.1f}GB '
          f'({(time.time()-t0)/60:.1f}min)', flush=True)
    return B, S, Y


def train(net, cache, epochs=15, bs=32, lr=1e-3, class_weight=None):
    B, S, Y = cache
    n = len(B); A = LatentAdapter()
    opt = torch.optim.Adam(A.parameters(), lr=lr)
    w = None if class_weight is None else torch.tensor(class_weight, dtype=torch.float32)
    lossf = nn.CrossEntropyLoss(weight=w)
    for p in net.parameters():
        p.requires_grad_(False)
    for ep in range(epochs):
        perm = np.random.permutation(n); tot = 0.0; t0 = time.time()
        for i in range(0, n, bs):
            idx = perm[i:i + bs]
            b = torch.from_numpy(B[idx].astype(np.float32))
            skips = tuple(torch.from_numpy(S[j][idx].astype(np.float32)) for j in range(4))
            y = torch.from_numpy(Y[idx].astype(np.int64))
            opt.zero_grad()
            logits = decode(net, A(b), skips)           # (bs,4,WIN)
            loss = lossf(logits, y)
            loss.backward(); opt.step()
            tot += loss.item() * len(idx)
        print(f'  ep{ep+1}/{epochs} loss={tot/n:.4f} ({time.time()-t0:.0f}s)', flush=True)
    return A.eval()


# ── dense inference with adapter at bottleneck (overlap-add, mirrors E.predict_dense) ──
def predict_dense_latent(net, A, sig1d, stride=512, batch=32):
    N = len(sig1d); L = max(N, WIN)
    starts = list(range(0, L - WIN + 1, stride))
    if not starts or starts[-1] != L - WIN:
        starts.append(L - WIN)
    psum = np.zeros((N_CLASS, L)); cnt = np.zeros(L); buf, bs = [], []

    def flush():
        nonlocal buf, bs
        if not buf:
            return
        x = torch.from_numpy(np.stack(buf))
        with torch.no_grad():
            b, skips = encode(net, x)
            logits = decode(net, A(b), skips)
            prob = torch.softmax(logits, dim=1).numpy()
        for bi, s0 in enumerate(bs):
            psum[:, s0:s0 + WIN] += prob[bi]; cnt[s0:s0 + WIN] += 1
        buf, bs = [], []

    for s0 in starts:
        buf.append(E.features_5(sig1d[s0:s0 + WIN])); bs.append(s0)
        if len(buf) >= batch:
            flush()
    flush()
    cnt[cnt == 0] = 1
    return np.argmax(psum / cnt, axis=0)[:N]


def heldout_signals():
    sigs = sorted(os.listdir(os.path.join(NPY, 'mixture')))
    v1done = set(json.load(open(os.path.join(EVAL_DIR, 'v1.json'))).keys())
    by = {'easy': [], 'medium': [], 'difficult': []}
    for s in sigs:
        if s in v1done:
            by[snr_cat(s)].append(s)
    return [by[c][i] for c in by for i in range(3)]   # 9, acelasi set ca decoded adapter


def evaluate(net, A, model, eval_sigs, stride=3840):
    """GT and base are read from the precomputed JSON (GT.json/v1.json, stride 3840,
    same macro-F1); only the LATENT prediction is computed fresh, at the same stride."""
    gtj = json.load(open(os.path.join(EVAL_DIR, 'GT.json')))
    bsj = json.load(open(os.path.join(EVAL_DIR, 'v1.json')))
    from config import FECG_ROOT as _r; inferred = _r + '/inferred_v1'
    rows = []
    for nm in eval_sigs:
        npy = os.path.join(inferred, nm + '.npy')           # = R.infer('v1') determinist
        ext = np.load(npy) if os.path.exists(npy) else R.infer('v1', model, load_sig('mixture', nm))
        mc = load_mask('mc_masks', nm).astype(int)
        N = min(ext.shape[1], len(mc)); mc = mc[:N]
        gch = gtj[nm]['channels']; bch = bsj[nm]['channels']
        a_lat, f_lat = [], []
        for ch in range(6):
            pl = predict_dense_latent(net, A, ext[ch, :N], stride=stride)[:N]
            a_lat.append(accuracy_score(mc, pl))
            f_lat.append(f1_score(mc, pl, average='macro', zero_division=0))
        row = dict(name=nm, cat=snr_cat(nm),
                   acc_gt=float(np.mean([c['m15_mc_acc'] for c in gch])),
                   acc_base=float(np.mean([c['m15_mc_acc'] for c in bch])),
                   acc_lat=float(np.mean(a_lat)),
                   f1_gt=float(np.mean([c['m15_mc_f1m'] for c in gch])),
                   f1_base=float(np.mean([c['m15_mc_f1m'] for c in bch])),
                   f1_lat=float(np.mean(f_lat)))
        rows.append(row)
        print(f"  {nm.split('_SNR')[0]:<12} {row['cat']:<10} "
              f"acc GT={row['acc_gt']:.3f} base={row['acc_base']:.3f} lat={row['acc_lat']:.3f} | "
              f"mF1 GT={row['f1_gt']:.3f} base={row['f1_base']:.3f} lat={row['f1_lat']:.3f}", flush=True)
    return rows


def write_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.load_workbook(XLSX); SH = 'Adapter latent v1'
    if SH in wb.sheetnames:
        del wb[SH]
    ws = wb.create_sheet(SH)
    bold = Font(bold=True); ital = Font(italic=True, size=9); ctr = Alignment(horizontal='center')
    thin = Side(style='thin', color='BBBBBB'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    L = [('ADAPTER LATENT (transfer la bottleneck-ul M15), antrenat pe DB_1, eval 9 held-out din test', bold),
         ('A = MLP per-pozitie pe bottleneck (256ch); loss = CrossEntropy prin decoderul INGHETAT vs masca categorii.', ital),
         ('LABEL-ONLY: nu se foloseste semnal GT nicaieri (nici la antrenare). Acc + macro-F1, 4 clase, medie pe 6 canale.', ital),
         ('GT = plafon (M15 pe fECG curat); base = M15 pe extras (necorectat); lat = M15 cu adapter latent.', ital), ('', None)]
    r = 1
    for txt, f in L:
        cc = ws.cell(r, 1, txt)
        if f:
            cc.font = f
        r += 1
    hdr = ['Semnal', 'Categorie', 'acc GT', 'acc base', 'acc latent', 'd acc',
           'mF1 GT', 'mF1 base', 'mF1 latent', 'd mF1']
    for j, h in enumerate(hdr):
        cc = ws.cell(r, 1 + j, h); cc.font = bold; cc.alignment = ctr
        cc.fill = PatternFill('solid', fgColor='FCE4D6'); cc.border = bd
    r += 1
    for x in rows:
        vals = [x['name'].split('_SNR')[0], x['cat'],
                round(x['acc_gt'], 3), round(x['acc_base'], 3), round(x['acc_lat'], 3), round(x['acc_lat'] - x['acc_base'], 3),
                round(x['f1_gt'], 3), round(x['f1_base'], 3), round(x['f1_lat'], 3), round(x['f1_lat'] - x['f1_base'], 3)]
        for j, v in enumerate(vals):
            cc = ws.cell(r, 1 + j, v); cc.border = bd
            if j >= 2:
                cc.alignment = ctr
        r += 1
    def m(k): return float(np.mean([x[k] for x in rows]))
    mv = ['MEDIE', '', round(m('acc_gt'), 3), round(m('acc_base'), 3), round(m('acc_lat'), 3), round(m('acc_lat') - m('acc_base'), 3),
          round(m('f1_gt'), 3), round(m('f1_base'), 3), round(m('f1_lat'), 3), round(m('f1_lat') - m('f1_base'), 3)]
    for j, v in enumerate(mv):
        cc = ws.cell(r, 1 + j, v); cc.font = bold; cc.border = bd
        if j >= 2:
            cc.alignment = ctr
    ws.column_dimensions['A'].width = 16
    for col in 'BCDEFGHIJ':
        ws.column_dimensions[col].width = 12
    wb.save(XLSX)
    print(f'\nfila "{SH}" scrisa in {XLSX}', flush=True)


def main():
    probe = 'probe' in sys.argv
    n_files = 4 if probe else 24
    n_windows = 600 if probe else 4000
    epochs = 2 if probe else 10

    model = R.load_extractor('v1')
    net = E.load_model(15)
    mats = sorted(glob.glob(os.path.join(DB1_LONG, '*.mat')))
    step = max(1, len(mats) // n_files)
    train_mats = mats[::step][:n_files]
    print(f'{"PROBE " if probe else ""}Training on {len(train_mats)} DB_1 .mat files, target {n_windows} windows.', flush=True)

    t = time.time()
    cache = build_cache(train_mats, model, net, n_windows)
    # class weights (inv-freq, tempered) to counter no-move class dominance
    yflat = cache[2].ravel(); freq = np.bincount(yflat, minlength=N_CLASS) + 1
    cw = (freq.sum() / freq); cw = (cw / cw.mean()) ** 0.5     # temperat
    print(f'  class freq {freq.tolist()}  -> weights {np.round(cw,2).tolist()}', flush=True)

    A = train(net, cache, epochs=epochs, class_weight=cw)
    print(f'  training done ({(time.time()-t)/60:.1f}min total)', flush=True)
    if probe:
        print('PROBE done (no eval/save).'); return

    torch.save(A.state_dict(), CKPT)
    rows = evaluate(net, A, model, heldout_signals())
    def m(k): return float(np.mean([x[k] for x in rows]))
    print(f'\n=== DOWNSTREAM M15 (adapter latent, 9 held-out) ===')
    print(f"  acc:    GT={m('acc_gt'):.3f}  base={m('acc_base'):.3f}  latent={m('acc_lat'):.3f}  d={m('acc_lat')-m('acc_base'):+.3f}")
    print(f"  macroF1:GT={m('f1_gt'):.3f}  base={m('f1_base'):.3f}  latent={m('f1_lat'):.3f}  d={m('f1_lat')-m('f1_base'):+.3f}")
    print(f"  (decoded adapter DB_1: acc 0.362->0.462, +0.100)")
    json.dump({'rows': rows, 'mean': {k: m(k) for k in
              ['acc_gt','acc_base','acc_lat','f1_gt','f1_base','f1_lat']}}, open(OUT, 'w'), indent=2)
    write_xlsx(rows)


if __name__ == '__main__':
    main()
