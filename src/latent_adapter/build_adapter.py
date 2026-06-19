"""
BEAT-LEVEL ADAPTER (POC) between the v1 extraction and the (frozen) M15 classifier.
Corrects the R-peak amplitude sequence of the extracted signal toward the GT form,
re-injects it as gain in the signal, then runs M15 on the corrected signal.

Leakage-free eval: train on signals disjoint from eval. The GT ceiling and the uncorrected
variant (v1) are read from GT.json/v1.json (the big job); only the corrected M15 is computed
here. Plus an envelope-level eval (per-class correlation + no-move variance).
"""
import sys, os, json, re, time
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
import pathsetup  # noqa
import numpy as np
import torch, torch.nn as nn
torch.set_num_threads(1)
from diag_info_ceiling import load_sig, load_mask, detect_peaks, amp_at
import pipeline_registry as R
import pretrained_clf as E
from sklearn.metrics import accuracy_score, f1_score

from config import NPY, EVAL_DIR, XLSX, RESULTS_DIR as _RES
OUT = _RES + '/adapter_v1_poc.json'
FS = 500; K = 31; HALF = K // 2; STRIDE = 3840; EPS = 1e-6
PATTERN = re.compile(r'SNRmn=([-\d]+)dB_SNRfm=([-\d]+)dB_SNRfn=([-\d]+)dB')
CLASS_NAME = {0:'no-move',1:'linear',2:'spline',3:'helix'}


def snr_cat(name):
    m = PATTERN.search(name); avg = (int(m.group(1))+int(m.group(2))+int(m.group(3)))/3
    return ('difficult' if avg < -5 else 'medium' if avg <= 5 else 'easy')


class Adapter(nn.Module):
    def __init__(self, k=K, h=64):
        super().__init__()
        self.net = nn.Sequential(nn.Linear(k, h), nn.ReLU(), nn.Linear(h, h), nn.ReLU(), nn.Linear(h, 1))
    def forward(self, x): return self.net(x).squeeze(-1)


def beat_seqs(name, ext_model):
    """-> per-channel list of (peaks, e_ext_norm, g_gt_norm, classes) at R-peak positions in EXTRACTED signal."""
    gt = load_sig('signals', name)
    ext = R.infer('v1', ext_model, load_sig('mixture', name))
    mc = load_mask('mc_masks', name).astype(int)
    N = min(gt.shape[1], ext.shape[1], len(mc))
    out = []
    for ch in range(6):
        pk = detect_peaks(ext[ch, :N])
        if len(pk) < K + 5:
            out.append(None); continue
        a_ext = amp_at(ext[ch, :N], pk); a_gt = amp_at(gt[ch, :N], pk)
        e = a_ext / (np.median(a_ext) + EPS)
        g = a_gt / (np.median(a_gt) + EPS)
        out.append((pk, e.astype(np.float32), g.astype(np.float32), mc[pk]))
    return ext, N, out


def windows(e):
    ep = np.pad(e, HALF, mode='edge')
    return np.stack([ep[i:i+K] for i in range(len(e))]).astype(np.float32)


def train_adapter(train_sigs, ext_model):
    X, Y = [], []
    for nm in train_sigs:
        _, _, seqs = beat_seqs(nm, ext_model)
        for s in seqs:
            if s is None: continue
            _, e, g, _ = s
            X.append(windows(e)); Y.append(g)
    X = np.concatenate(X); Y = np.concatenate(Y)
    print(f'  train pairs: {len(X)}', flush=True)
    Xt = torch.from_numpy(X); Yt = torch.from_numpy(Y)
    m = Adapter(); opt = torch.optim.Adam(m.parameters(), lr=1e-3); lossf = nn.MSELoss()
    n = len(Xt); bs = 512
    for ep in range(40):
        perm = torch.randperm(n)
        for i in range(0, n, bs):
            idx = perm[i:i+bs]; opt.zero_grad()
            loss = lossf(m(Xt[idx]), Yt[idx]); loss.backward(); opt.step()
    m.eval()
    return m


def correct_signal(x, pk, e, adapter):
    with torch.no_grad():
        g_hat = adapter(torch.from_numpy(windows(e))).numpy()
    gain_pk = np.clip(g_hat / (e + EPS), 0.25, 4.0)
    gain = np.interp(np.arange(len(x)), pk, gain_pk, left=gain_pk[0], right=gain_pk[-1])
    return (x * gain).astype(np.float32)


def classify_m15(sig, clf15):
    L = len(sig); starts = list(range(0, L - 3840 + 1, STRIDE))
    if starts[-1] != L - 3840: starts.append(L - 3840)
    f5 = np.stack([E.features_5(sig[s:s+3840]) for s in starts]).astype(np.float32)
    psum = np.zeros((4, L)); cnt = np.zeros(L)
    with torch.no_grad():
        for i in range(0, len(f5), 64):
            p = torch.softmax(clf15(torch.from_numpy(f5[i:i+64])), dim=1).numpy()
            for bi in range(p.shape[0]):
                s0 = starts[i+bi]; psum[:, s0:s0+3840] += p[bi]; cnt[s0:s0+3840] += 1
    cnt[cnt == 0] = 1
    return np.argmax(psum/cnt, axis=0)[:L]


def env_corr(seqs_eval):
    """Per-class envelope correlation: uncorrected(e) vs GT(g) and corrected(g_hat) vs GT, + no-move std."""
    res = {c: {'unc': [], 'cor': []} for c in (0,1,2,3)}
    nm_std = {'gt': [], 'unc': [], 'cor': []}
    for pk, e, g, cls, g_hat in seqs_eval:
        b = 0
        while b < len(pk):
            ed = b
            while ed+1 < len(pk) and cls[ed+1] == cls[b]: ed += 1
            run = slice(b, ed+1); L = ed+1-b; c = int(cls[b])
            if L >= 6 and g[run].std() > 1e-6:
                if e[run].std() > 1e-6: res[c]['unc'].append(np.corrcoef(g[run], e[run])[0,1])
                if g_hat[run].std() > 1e-6: res[c]['cor'].append(np.corrcoef(g[run], g_hat[run])[0,1])
            if c == 0 and L >= 6:
                nm_std['gt'].append(g[run].std()); nm_std['unc'].append(e[run].std()); nm_std['cor'].append(g_hat[run].std())
            b = ed+1
    return res, nm_std


def main():
    sigs = sorted(os.listdir(os.path.join(NPY, 'mixture')))
    v1done = set(json.load(open(os.path.join(EVAL_DIR,'v1.json'))).keys())
    by = {'easy': [], 'medium': [], 'difficult': []}
    for s in sigs:
        if s in v1done: by[snr_cat(s)].append(s)
    eval_sigs = [by[c][i] for c in by for i in range(3)]          # 9 eval (3/cat)
    train_sigs = [s for c in by for s in by[c][3:11]]             # ~24 train (8/cat)
    print(f'train={len(train_sigs)} eval={len(eval_sigs)}', flush=True)

    ext_model = R.load_extractor('v1')
    t = time.time()
    adapter = train_adapter(train_sigs, ext_model)
    print(f'  adapter trained in {(time.time()-t)/60:.1f}min', flush=True)

    clf15 = E.load_model(15)
    gt_json = json.load(open(os.path.join(EVAL_DIR,'GT.json')))
    v1_json = json.load(open(os.path.join(EVAL_DIR,'v1.json')))

    rows = []; seqs_eval_all = []
    for nm in eval_sigs:
        ext, N, seqs = beat_seqs(nm, ext_model)
        mc = load_mask('mc_masks', nm).astype(int)[:N]
        cat = snr_cat(nm)
        acc_cor = []
        for ch in range(6):
            if seqs[ch] is None: continue
            pk, e, g, cls = seqs[ch]
            with torch.no_grad():
                g_hat = adapter(torch.from_numpy(windows(e))).numpy()
            seqs_eval_all.append((pk, e, g, cls, g_hat))
            xc = correct_signal(ext[ch, :N], pk, e, adapter)
            p15 = classify_m15(xc, clf15)[:N]
            acc_cor.append(accuracy_score(mc, p15))
        # GT ceiling + uncorrected v1 from JSON
        gt_acc = np.mean([c['m15_mc_acc'] for c in gt_json[nm]['channels']])
        v1_acc = np.mean([c['m15_mc_acc'] for c in v1_json[nm]['channels']])
        cor_acc = float(np.mean(acc_cor))
        rows.append((nm, cat, gt_acc, v1_acc, cor_acc))
        print(f'  {nm.split("_SNR")[0]:<12} {cat:<10} GT={gt_acc:.3f}  v1={v1_acc:.3f}  v1+adapter={cor_acc:.3f}', flush=True)

    ec, nm_std = env_corr(seqs_eval_all)
    print('\n=== ENVELOPE (per-class correlation with GT) ===')
    for c in (1,2,3):
        u = np.nanmean(ec[c]['unc']) if ec[c]['unc'] else float('nan')
        co = np.nanmean(ec[c]['cor']) if ec[c]['cor'] else float('nan')
        print(f'  {CLASS_NAME[c]:<8} uncorrected r={u:.3f} -> corrected r={co:.3f}')
    print(f'  no-move std (smaller=flatter): GT={np.nanmean(nm_std["gt"]):.3f} '
          f'uncorrected={np.nanmean(nm_std["unc"]):.3f} corrected={np.nanmean(nm_std["cor"]):.3f}')

    print('\n=== DOWNSTREAM M15 acc (mean over channels) ===')
    g = np.mean([r[2] for r in rows]); v = np.mean([r[3] for r in rows]); c = np.mean([r[4] for r in rows])
    print(f'  GT={g:.3f}  v1(uncorrected)={v:.3f}  v1+adapter={c:.3f}  | delta={c-v:+.3f}')
    json.dump({'rows': rows, 'mean': {'GT': g, 'v1': v, 'v1_adapter': c}}, open(OUT, 'w'), indent=2)
    write_xlsx(rows, ec, nm_std)


def write_xlsx(rows, ec, nm_std):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.load_workbook(XLSX); SH = 'Adapter v1 POC'
    if SH in wb.sheetnames: del wb[SH]
    ws = wb.create_sheet(SH)
    bold = Font(bold=True); ital = Font(italic=True, size=9); ctr = Alignment(horizontal='center')
    thin = Side(style='thin', color='BBBBBB'); bd = Border(left=thin,right=thin,top=thin,bottom=thin)
    L = [('Beat-level ADAPTER v1 -> M15 (POC, eval on held-out signals)', bold),
         ('Corrects the R-peak amplitude sequence of the v1 extracted signal toward GT form (MLP with K=31 beat context), re-injects as gain, then frozen M15.', ital),
         ('GT = ceiling (clean signal). v1 = M15 on uncorrected extracted. v1+adapter = M15 on corrected extracted. All = mean over 6 channels, 4-class multiclass acc.', ital),
         ('Train/eval disjoint (no leakage). GT and v1 read from precomputed JSON; only v1+adapter computed here.', ital), ('', None)]
    r = 1
    for txt, f in L:
        cc = ws.cell(r,1,txt)
        if f: cc.font = f
        r += 1
    for j,h in enumerate(['Signal','Category','GT (ceiling)','v1 (uncorrected)','v1+adapter','delta']):
        cc = ws.cell(r,1+j,h); cc.font = bold; cc.alignment = ctr; cc.fill = PatternFill('solid',fgColor='DDEBF7'); cc.border = bd
    r += 1
    for nm, cat, g, v, c in rows:
        vals = [nm.split('_SNR')[0], cat, round(g,3), round(v,3), round(c,3), round(c-v,3)]
        for j,val in enumerate(vals):
            cc = ws.cell(r,1+j,val); cc.border = bd
            if j >= 2: cc.alignment = ctr
        r += 1
    gm = np.mean([x[2] for x in rows]); vm = np.mean([x[3] for x in rows]); cm = np.mean([x[4] for x in rows])
    for j,val in enumerate(['MEAN','',round(gm,3),round(vm,3),round(cm,3),round(cm-vm,3)]):
        cc = ws.cell(r,1+j,val); cc.font = bold; cc.border = bd
        if j >= 2: cc.alignment = ctr
    r += 2
    ws.cell(r,1,'ENVELOPE — correlation with GT (per class), uncorrected -> corrected:').font = bold; r += 1
    for cc_ in (1,2,3):
        u = np.nanmean(ec[cc_]['unc']) if ec[cc_]['unc'] else float('nan')
        co = np.nanmean(ec[cc_]['cor']) if ec[cc_]['cor'] else float('nan')
        ws.cell(r,1,f'  {CLASS_NAME[cc_]}: r {u:.3f} -> {co:.3f}'); r += 1
    ws.cell(r,1,f'  no-move std (smaller=flatter): GT={np.nanmean(nm_std["gt"]):.3f} uncor={np.nanmean(nm_std["unc"]):.3f} cor={np.nanmean(nm_std["cor"]):.3f}'); r += 1
    ws.column_dimensions['A'].width = 16
    for col in 'BCDEF': ws.column_dimensions[col].width = 14
    wb.save(XLSX)
    print(f'\nsheet "{SH}" written to {XLSX}')


if __name__ == '__main__':
    main()
