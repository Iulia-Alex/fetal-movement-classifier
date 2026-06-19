"""
BINARY METRICS (movement yes/no) for the latent adapter, alongside the multiclass macro-F1.
M15 is multiclass -> binary = (argmax > 0). Comparable to Edward's '0.91' figure
(which is binary ACCURACY on the best channel).

Two sets:
  HELD-OUT (9, Final_Test_DB_npy) — the strictly disjoint set used for adapter training
  TEST_DB  (11, Sem1..Sem11)      — a fully disjoint set (generalization)

Per signal, per channel, stride 3840: GT / base(v1) / latent. Reports:
  multiclass: acc, macro-F1   |   binary: acc, F1
aggregation: mean over 6 channels AND the best channel (by GT binary F1).
Caches predictions in .npz (instant re-run). xlsx 'Adapter binar' + json.
"""
import sys, os, glob, time, json
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
import pathsetup  # noqa
import numpy as np
import torch
os.environ.setdefault('LA_THREADS', '12')
torch.set_num_threads(int(os.environ['LA_THREADS']))
import scipy.io as sio
from sklearn.metrics import accuracy_score, f1_score

import pipeline_registry as R
import pretrained_clf as E
from latent_adapter import LatentAdapter, predict_dense_latent, heldout_signals
from diag_info_ceiling import load_sig, load_mask
from movement_dataset import _extract_fecg

from config import XLSX, TEST_DIR, RESULTS_DIR as _RES, FECG_ROOT as _root
CKPT     = _RES   + '/adapter_latent_v1.pt'
OUT      = _RES   + '/adapter_binary.json'
CACHE    = _RES   + '/adapter_binary_preds.npz'
INFERRED = _root  + '/inferred_v1'
STRIDE = 3840
SEMS = [f'Sem{i}' for i in range(1, 12)]

_pred_cache = dict(np.load(CACHE, allow_pickle=True)) if os.path.exists(CACHE) else {}


def get_preds(key, signal_fn):
    """signal_fn() -> (fecg(6,N), ext(6,N), mc(N)); returns per-channel p_gt/p_base/p_lat (6,N)."""
    if key in _pred_cache:
        d = _pred_cache[key].item()
        return d['mc'], d['gt'], d['base'], d['lat']
    fecg, ext, mc = signal_fn()
    N = min(ext.shape[1], len(mc), fecg.shape[1]); mc = mc[:N]
    pg, pb, pl = [], [], []
    for ch in range(6):
        pg.append(E.run_model(15, fecg[ch, :N], NET, stride=STRIDE)[0][:N])
        pb.append(E.run_model(15, ext[ch, :N], NET, stride=STRIDE)[0][:N])
        pl.append(predict_dense_latent(NET, ADP, ext[ch, :N], stride=STRIDE)[:N])
    gt = np.stack(pg).astype(np.int8); base = np.stack(pb).astype(np.int8); lat = np.stack(pl).astype(np.int8)
    _pred_cache[key] = np.array({'mc': mc.astype(np.int8), 'gt': gt, 'base': base, 'lat': lat}, dtype=object)
    np.savez(CACHE, **_pred_cache)
    return mc, gt, base, lat


def metrics(mc, pred6):
    """Per-channel (mc_acc, mc_f1m, bin_acc, bin_f1) -> dict with mean + best-channel."""
    mb = (mc > 0).astype(int)
    rows = []
    for ch in range(6):
        p = pred6[ch].astype(int); pb = (p > 0).astype(int)
        rows.append((accuracy_score(mc, p), f1_score(mc, p, average='macro', zero_division=0),
                     accuracy_score(mb, pb), f1_score(mb, pb, zero_division=0)))
    rows = np.array(rows)  # (6,4)
    return rows


def aggregate(stem, mc, gt, base, lat):
    M = {k: metrics(mc, v) for k, v in (('gt', gt), ('base', base), ('lat', lat))}
    bestch = int(np.argmax(M['gt'][:, 3]))   # channel with the best GT binary F1
    out = {'name': stem, 'best_ch': bestch}
    cols = ['mc_acc', 'mc_f1m', 'bin_acc', 'bin_f1']
    for k in ('gt', 'base', 'lat'):
        for ci, c in enumerate(cols):
            out[f'{k}_{c}_mean'] = float(M[k][:, ci].mean())
            out[f'{k}_{c}_best'] = float(M[k][bestch, ci])
    return out


def sem_fn(stem):
    def f():
        mat = sio.loadmat(os.path.join(TEST_DIR, stem + '.mat')); o = mat['out']
        mix = o['mixture'][0][0].astype(np.float32); fecg = _extract_fecg(o)
        if fecg.shape[0] < mix.shape[0]:
            fecg = np.repeat(fecg, mix.shape[0] // fecg.shape[0], axis=0)
        cat = o['category_mask'][0][0].ravel().astype(int); del mat
        return fecg, R.infer('v1', V1, mix), cat
    return f


def heldout_fn(nm):
    def f():
        fecg = load_sig('signals', nm)
        npy = os.path.join(INFERRED, nm + '.npy')
        ext = np.load(npy) if os.path.exists(npy) else R.infer('v1', V1, load_sig('mixture', nm))
        mc = load_mask('mc_masks', nm).astype(int)
        return fecg, ext, mc
    return f


def run_set(label, items, fn_factory, t0):
    rows = []
    for it in items:
        stem = it.split('_SNR')[0]
        mc, gt, base, lat = get_preds(f'{label}:{it}', fn_factory(it))
        r = aggregate(stem, mc, gt, base, lat); r['set'] = label
        rows.append(r)
        print(f"  [{label}] {stem:<12} mF1 GT={r['gt_mc_f1m_mean']:.3f} base={r['base_mc_f1m_mean']:.3f} "
              f"lat={r['lat_mc_f1m_mean']:.3f} | binF1 base={r['base_bin_f1_mean']:.3f} lat={r['lat_bin_f1_mean']:.3f}"
              f" | binACC best GT={r['gt_bin_acc_best']:.3f} lat={r['lat_bin_acc_best']:.3f}"
              f"  [{(time.time()-t0)/60:.1f}min]", flush=True)
    return rows


def write_xlsx(allrows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.load_workbook(XLSX); SH = 'Adapter binar'
    if SH in wb.sheetnames:
        del wb[SH]
    ws = wb.create_sheet(SH)
    bold = Font(bold=True); ital = Font(italic=True, size=9); ctr = Alignment(horizontal='center')
    thin = Side(style='thin', color='BBBBBB'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    L = [('LATENT ADAPTER — BINARY metrics (movement yes/no, M15: class>0) alongside the multiclass macro-F1', bold),
         ('Comparison with Edward\'s "0.91" figure = binary ACC on the BEST channel. Here: mean-over-6-channels AND best-channel.', ital),
         ('GT=M15 on clean fECG (ceiling); base=M15 on extracted v1; latent=M15+latent adapter. HELD-OUT=9 disjoint; TEST_DB=Sem1-11.', ital),
         ('', None)]
    r = 1
    for txt, f in L:
        cc = ws.cell(r, 1, txt)
        if f:
            cc.font = f
        r += 1
    hdr = ['Set', 'Signal', 'mF1 GT', 'mF1 base', 'mF1 lat', 'binF1 GT', 'binF1 base', 'binF1 lat',
           'binACC GT (best)', 'binACC base (best)', 'binACC lat (best)']
    keys = [('gt_mc_f1m_mean',), ('base_mc_f1m_mean',), ('lat_mc_f1m_mean',),
            ('gt_bin_f1_mean',), ('base_bin_f1_mean',), ('lat_bin_f1_mean',),
            ('gt_bin_acc_best',), ('base_bin_acc_best',), ('lat_bin_acc_best',)]
    for j, h in enumerate(hdr):
        cc = ws.cell(r, 1 + j, h); cc.font = bold; cc.alignment = ctr
        cc.fill = PatternFill('solid', fgColor='DDEBF7'); cc.border = bd
    r += 1

    def emit(row, vals, f=None):
        for j, v in enumerate(vals):
            cc = ws.cell(row, 1 + j, v); cc.border = bd
            if f:
                cc.font = f
            if j >= 2:
                cc.alignment = ctr
    for x in allrows:
        emit(r, [x['set'], x['name']] + [round(x[k[0]], 3) for k in keys]); r += 1
    # means per set
    for label in ['HELD-OUT', 'TEST_DB']:
        sub = [x for x in allrows if x['set'] == label]
        if not sub:
            continue
        mv = [label, 'MEAN'] + [round(float(np.mean([x[k[0]] for x in sub])), 3) for k in keys]
        emit(r, mv, f=bold); r += 1
    ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 13
    for col in 'CDEFGHIJK':
        ws.column_dimensions[col].width = 13
    wb.save(XLSX)
    print(f'\nsheet "{SH}" written to {XLSX}', flush=True)


def main():
    global V1, NET, ADP
    print(f'Loading v1 + M15 + latent adapter...', flush=True)
    V1 = R.load_extractor('v1'); NET = E.load_model(15)
    ADP = LatentAdapter(); ADP.load_state_dict(torch.load(CKPT, map_location='cpu')); ADP.eval()
    t0 = time.time()
    allrows = []
    print('=== HELD-OUT (9 disjuncte) ===', flush=True)
    allrows += run_set('HELD-OUT', heldout_signals(), heldout_fn, t0)
    print('=== TEST_DB (Sem1..Sem11) ===', flush=True)
    allrows += run_set('TEST_DB', SEMS, sem_fn, t0)

    def setmean(label, k):
        sub = [x for x in allrows if x['set'] == label]
        return float(np.mean([x[k] for x in sub]))
    print('\n=== SUMMARY (mean over channels) ===', flush=True)
    for label in ['HELD-OUT', 'TEST_DB']:
        print(f'  {label}: mF1 GT={setmean(label,"gt_mc_f1m_mean"):.3f} base={setmean(label,"base_mc_f1m_mean"):.3f} '
              f'lat={setmean(label,"lat_mc_f1m_mean"):.3f}  ||  binF1 GT={setmean(label,"gt_bin_f1_mean"):.3f} '
              f'base={setmean(label,"base_bin_f1_mean"):.3f} lat={setmean(label,"lat_bin_f1_mean"):.3f}', flush=True)
        print(f'         binACC best-channel: GT={setmean(label,"gt_bin_acc_best"):.3f} '
              f'base={setmean(label,"base_bin_acc_best"):.3f} lat={setmean(label,"lat_bin_acc_best"):.3f}', flush=True)
    json.dump(allrows, open(OUT, 'w'), indent=2)
    write_xlsx(allrows)
    print(f'\nGata ({(time.time()-t0)/60:.1f}min).', flush=True)


if __name__ == '__main__':
    main()
